"""
Generador de informe fiscal IRPF España para criptomonedas.

Aplica el método FIFO obligatorio según la AEAT para el cálculo
de ganancias y pérdidas patrimoniales derivadas de transmisiones
de criptomonedas (base del ahorro, Modelo 100 casilla 0389).

Nota DAC8: desde 2026 los exchanges están obligados a reportar
automáticamente a la AEAT. Este informe es para verificación propia
y para rellenar el Modelo 100 / Anexo H con seguridad.
"""
from __future__ import annotations

import json
from collections import deque
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, ROUND_DOWN, ROUND_HALF_UP
from pathlib import Path
from typing import NamedTuple
from zoneinfo import ZoneInfo

from loguru import logger
from sqlalchemy.orm import Session

from core.database import Trade, get_trades

# ---------------------------------------------------------------------------
# Constantes IRPF 2026 (declaración del año fiscal 2025)
# Fuente: Ley 35/2006 del IRPF, artículos 66-66bis
# ---------------------------------------------------------------------------

MADRID_TZ = ZoneInfo("Europe/Madrid")

# Tramos de la base del ahorro: (límite acumulado, tipo marginal)
# None como límite = tramo sin techo
_IRPF_BRACKETS: list[tuple[Decimal | None, Decimal]] = [
    (Decimal("6000"),   Decimal("0.19")),
    (Decimal("50000"),  Decimal("0.21")),
    (Decimal("200000"), Decimal("0.23")),
    (None,              Decimal("0.28")),
]

_REPORTS_DIR = Path(__file__).parent.parent / "reports"


# ---------------------------------------------------------------------------
# Estructuras de datos
# ---------------------------------------------------------------------------

@dataclass
class PurchaseLot:
    """Lote de compra pendiente de emparejarse con una venta (FIFO)."""
    trade_id: int
    symbol: str
    date: datetime           # UTC
    quantity: Decimal        # moneda base (e.g. BTC)
    unit_price_eur: Decimal  # EUR por unidad
    fee_eur: Decimal         # comisión total en EUR (ya incluida en cost basis)

    @property
    def total_cost_eur(self) -> Decimal:
        return self.quantity * self.unit_price_eur + self.fee_eur


@dataclass
class GainLossRecord:
    """Par compra-venta emparejados por FIFO."""
    symbol: str
    buy_date: datetime
    sell_date: datetime
    buy_price_eur: Decimal
    sell_price_eur: Decimal
    quantity: Decimal
    gross_gain_eur: Decimal
    total_fees_eur: Decimal
    net_gain_eur: Decimal    # positivo = ganancia, negativo = pérdida


class FiscalSummary(NamedTuple):
    year: int
    total_trades: int
    total_gains_eur: Decimal
    total_losses_eur: Decimal
    net_result_eur: Decimal
    estimated_tax_eur: Decimal
    deductible_fees_eur: Decimal
    carryover_losses_eur: Decimal
    generated_at: datetime


# ---------------------------------------------------------------------------
# Lógica FIFO
# ---------------------------------------------------------------------------

class FIFOCalculator:
    """
    Aplica el método FIFO para emparejar compras y ventas.

    Para cada símbolo mantiene una deque de lotes de compra.
    Cuando procesa una venta, consume los lotes más antiguos primero.
    """

    def __init__(self, usd_eur_rate: Decimal) -> None:
        self._rate = usd_eur_rate
        self._lots: dict[str, deque[PurchaseLot]] = {}
        self.gain_loss_records: list[GainLossRecord] = []
        self.unmatched_sells: list[Trade] = []

    def _to_eur(self, usdt_amount: Decimal) -> Decimal:
        return (usdt_amount * self._rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def process_trade(self, trade: Trade) -> None:
        """Procesa un trade y actualiza el estado FIFO."""
        if trade.side == "buy":
            self._add_lot(trade)
        elif trade.side == "sell":
            self._process_sell(trade)

    def _add_lot(self, trade: Trade) -> None:
        symbol = trade.symbol
        if symbol not in self._lots:
            self._lots[symbol] = deque()

        fee_eur = self._to_eur(trade.fee if trade.fee else Decimal("0"))
        unit_price_eur = self._to_eur(trade.price)

        self._lots[symbol].append(PurchaseLot(
            trade_id=trade.id,
            symbol=symbol,
            date=trade.timestamp,
            quantity=trade.quantity,
            unit_price_eur=unit_price_eur,
            fee_eur=fee_eur,
        ))

    def _process_sell(self, trade: Trade) -> None:
        symbol = trade.symbol
        lots = self._lots.get(symbol)
        if not lots:
            logger.warning(
                "FIFO: venta de {} sin lotes de compra disponibles (trade_id={})",
                symbol, trade.id,
            )
            self.unmatched_sells.append(trade)
            return

        remaining_qty = trade.quantity
        sell_price_eur = self._to_eur(trade.price)
        sell_fee_eur = self._to_eur(trade.fee if trade.fee else Decimal("0"))

        while remaining_qty > Decimal("0") and lots:
            lot = lots[0]

            if lot.quantity <= remaining_qty:
                # Consumes el lote entero
                matched_qty = lot.quantity
                # Fee de venta proporcional a la cantidad de este lote
                proportional_sell_fee = (sell_fee_eur * matched_qty / trade.quantity).quantize(
                    Decimal("0.00000001")
                )
                gross = (sell_price_eur - lot.unit_price_eur) * matched_qty
                total_fees = lot.fee_eur + proportional_sell_fee
                net = gross - total_fees

                self.gain_loss_records.append(GainLossRecord(
                    symbol=symbol,
                    buy_date=lot.date,
                    sell_date=trade.timestamp,
                    buy_price_eur=lot.unit_price_eur,
                    sell_price_eur=sell_price_eur,
                    quantity=matched_qty,
                    gross_gain_eur=gross.quantize(Decimal("0.01")),
                    total_fees_eur=total_fees.quantize(Decimal("0.01")),
                    net_gain_eur=net.quantize(Decimal("0.01")),
                ))
                remaining_qty -= lot.quantity
                lots.popleft()

            else:
                # Consume solo parte del lote
                matched_qty = remaining_qty
                proportional_buy_fee = (lot.fee_eur * matched_qty / lot.quantity).quantize(
                    Decimal("0.00000001")
                )
                proportional_sell_fee = (sell_fee_eur * matched_qty / trade.quantity).quantize(
                    Decimal("0.00000001")
                )
                gross = (sell_price_eur - lot.unit_price_eur) * matched_qty
                total_fees = proportional_buy_fee + proportional_sell_fee
                net = gross - total_fees

                self.gain_loss_records.append(GainLossRecord(
                    symbol=symbol,
                    buy_date=lot.date,
                    sell_date=trade.timestamp,
                    buy_price_eur=lot.unit_price_eur,
                    sell_price_eur=sell_price_eur,
                    quantity=matched_qty,
                    gross_gain_eur=gross.quantize(Decimal("0.01")),
                    total_fees_eur=total_fees.quantize(Decimal("0.01")),
                    net_gain_eur=net.quantize(Decimal("0.01")),
                ))
                # Actualizar el lote restante
                lot.fee_eur -= proportional_buy_fee
                lot.quantity -= matched_qty
                remaining_qty = Decimal("0")


# ---------------------------------------------------------------------------
# Cálculo IRPF progresivo
# ---------------------------------------------------------------------------

def calculate_irpf_tax(net_gain_eur: Decimal) -> Decimal:
    """
    Aplica los tramos progresivos de la base del ahorro IRPF 2026.
    Retorna la estimación del impuesto total. Retorna 0 si la ganancia es ≤ 0.
    """
    if net_gain_eur <= Decimal("0"):
        return Decimal("0")

    tax = Decimal("0")
    remaining = net_gain_eur
    prev_limit = Decimal("0")

    for limit, rate in _IRPF_BRACKETS:
        if limit is None:
            tax += remaining * rate
            break
        bracket_size = limit - prev_limit
        taxable_in_bracket = min(remaining, bracket_size)
        tax += taxable_in_bracket * rate
        remaining -= taxable_in_bracket
        prev_limit = limit
        if remaining <= Decimal("0"):
            break

    return tax.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


# ---------------------------------------------------------------------------
# Generador principal
# ---------------------------------------------------------------------------

class FiscalReportGenerator:
    def __init__(
        self,
        session: Session,
        usd_eur_rate: Decimal = Decimal("0.92"),
        carryover_losses_eur: Decimal = Decimal("0"),
    ) -> None:
        self._session = session
        self._rate = usd_eur_rate
        self._carryover = carryover_losses_eur
        _REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    def generate_annual_report(self, year: int) -> str:
        """
        Genera el informe IRPF para el año indicado.
        Retorna la ruta del archivo Excel generado.
        """
        logger.info("Generando informe fiscal IRPF {} (tasa {}/€)", year, self._rate)

        # 1 — Cargar todos los trades del año (paper=False para producción)
        from_dt = datetime(year, 1, 1, tzinfo=timezone.utc)
        to_dt = datetime(year, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
        trades = get_trades(self._session, from_dt=from_dt, to_dt=to_dt, is_paper=False)

        # También incluir trades paper si no hay trades reales (modo desarrollo)
        if not trades:
            trades = get_trades(self._session, from_dt=from_dt, to_dt=to_dt)
            if trades:
                logger.warning("Solo se encontraron trades en modo paper — informe fiscal de simulación")

        logger.info("Trades del año {}: {}", year, len(trades))

        # 2 — Aplicar FIFO
        fifo = FIFOCalculator(usd_eur_rate=self._rate)
        for trade in sorted(trades, key=lambda t: t.timestamp):
            fifo.process_trade(trade)

        # 3 — Calcular totales
        all_gains = [r.net_gain_eur for r in fifo.gain_loss_records if r.net_gain_eur > 0]
        all_losses = [r.net_gain_eur for r in fifo.gain_loss_records if r.net_gain_eur < 0]
        total_gains = sum(all_gains, Decimal("0"))
        total_losses = sum(all_losses, Decimal("0"))
        total_fees = sum((r.total_fees_eur for r in fifo.gain_loss_records), Decimal("0"))

        # Las pérdidas del año reducen las ganancias; el resto arrastra 4 años
        net_result = total_gains + total_losses  # total_losses ya es negativo
        effective_net = net_result + self._carryover  # aplicar pérdidas arrastradas

        estimated_tax = calculate_irpf_tax(effective_net)
        remaining_carryover = min(Decimal("0"), effective_net)  # solo si aún hay pérdidas

        summary = FiscalSummary(
            year=year,
            total_trades=len(trades),
            total_gains_eur=total_gains,
            total_losses_eur=total_losses,
            net_result_eur=net_result,
            estimated_tax_eur=estimated_tax,
            deductible_fees_eur=total_fees,
            carryover_losses_eur=abs(remaining_carryover),
            generated_at=datetime.now(timezone.utc),
        )

        # 4 — Generar archivos
        excel_path = self._write_excel(year, trades, fifo.gain_loss_records, summary)
        self._write_json(year, summary)

        logger.info(
            "Informe {} generado: {} trades | ganancia neta {:.2f}€ | impuesto estimado {:.2f}€",
            year, len(trades), net_result, estimated_tax,
        )
        return excel_path

    # -----------------------------------------------------------------------
    # Excel
    # -----------------------------------------------------------------------

    def _write_excel(
        self,
        year: int,
        trades: list[Trade],
        gl_records: list[GainLossRecord],
        summary: FiscalSummary,
    ) -> str:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            logger.error("openpyxl no instalado — solo se generará el JSON")
            return ""

        wb = Workbook()
        wb.remove(wb.active)  # eliminar hoja por defecto

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")
        alt_fill = PatternFill("solid", fgColor="D6E4F0")

        def style_header(ws, headers):
            ws.append(headers)
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

        def autofit(ws):
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        def fmt_date(dt: datetime) -> str:
            return dt.astimezone(MADRID_TZ).strftime("%d/%m/%Y %H:%M")

        def fmt_dec(v: Decimal, decimals: int = 2) -> str:
            fmt = Decimal("0." + "0" * decimals)
            return str(v.quantize(fmt, rounding=ROUND_HALF_UP))

        # ---- Pestaña 1: Operaciones ----
        ws1 = wb.create_sheet("Operaciones")
        style_header(ws1, [
            "Fecha", "Par", "Tipo", "Cantidad", "Precio (USDT)",
            "Total EUR", "Comisión", "Estrategia", "ID Orden", "Paper",
        ])
        for i, t in enumerate(sorted(trades, key=lambda x: x.timestamp), start=2):
            total_eur = self._rate * t.price * t.quantity
            ws1.append([
                fmt_date(t.timestamp),
                t.symbol,
                t.side.upper(),
                fmt_dec(t.quantity, 8),
                fmt_dec(t.price, 2),
                fmt_dec(total_eur, 2),
                fmt_dec(t.fee, 4),
                t.strategy,
                t.order_id,
                "Sí" if t.is_paper else "No",
            ])
            if i % 2 == 0:
                for cell in ws1[i]:
                    cell.fill = alt_fill
        autofit(ws1)

        # ---- Pestaña 2: Ganancias y Pérdidas ----
        ws2 = wb.create_sheet("Ganancias y Pérdidas")
        style_header(ws2, [
            "Fecha Cierre", "Activo", "Fecha Compra (FIFO)",
            "Precio Compra (€)", "Precio Venta (€)", "Cantidad",
            "Ganancia Bruta (€)", "Comisiones (€)", "Ganancia Neta (€)",
        ])
        for i, r in enumerate(gl_records, start=2):
            ws2.append([
                fmt_date(r.sell_date),
                r.symbol.split("-")[0],
                fmt_date(r.buy_date),
                fmt_dec(r.buy_price_eur),
                fmt_dec(r.sell_price_eur),
                fmt_dec(r.quantity, 8),
                fmt_dec(r.gross_gain_eur),
                fmt_dec(r.total_fees_eur),
                fmt_dec(r.net_gain_eur),
            ])
            if i % 2 == 0:
                for cell in ws2[i]:
                    cell.fill = alt_fill
            # Colorear ganancia neta
            net_cell = ws2.cell(row=i, column=9)
            net_cell.font = Font(
                color="006100" if r.net_gain_eur >= 0 else "9C0006",
                bold=True,
            )
        autofit(ws2)

        # ---- Pestaña 3: Resumen Fiscal ----
        ws3 = wb.create_sheet("Resumen Fiscal")
        ws3.column_dimensions["A"].width = 45
        ws3.column_dimensions["B"].width = 20

        header_row = ["Concepto", "Importe (€)"]
        ws3.append(header_row)
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill

        rows = [
            ("Año fiscal", str(summary.year)),
            ("Total operaciones registradas", str(summary.total_trades)),
            ("", ""),
            ("GANANCIAS", ""),
            ("Total ganancias realizadas", fmt_dec(summary.total_gains_eur)),
            ("Total pérdidas realizadas", fmt_dec(summary.total_losses_eur)),
            ("Resultado neto base ahorro", fmt_dec(summary.net_result_eur)),
            ("Pérdidas de años anteriores aplicadas", fmt_dec(self._carryover)),
            ("Base imponible del ahorro (efectiva)", fmt_dec(summary.net_result_eur + self._carryover)),
            ("", ""),
            ("IMPUESTO ESTIMADO", ""),
            ("Tramo hasta 6.000€ (19%)", ""),
            ("Tramo 6.000€ - 50.000€ (21%)", ""),
            ("Tramo 50.000€ - 200.000€ (23%)", ""),
            ("Tramo > 200.000€ (28%)", ""),
            ("Estimación IRPF total", fmt_dec(summary.estimated_tax_eur)),
            ("", ""),
            ("OTROS", ""),
            ("Comisiones totales deducibles", fmt_dec(summary.deductible_fees_eur)),
            ("Pérdidas arrastrables (4 años)", fmt_dec(summary.carryover_losses_eur)),
            ("", ""),
            ("Informe generado", fmt_date(summary.generated_at)),
            ("Tipo de cambio USDT/EUR aplicado", str(self._rate)),
        ]
        for row_data in rows:
            ws3.append(list(row_data))

        # Resaltar filas clave
        result_row = 8  # "Resultado neto base ahorro"
        ws3.cell(row=result_row, column=1).font = Font(bold=True)
        ws3.cell(row=result_row, column=2).font = Font(
            bold=True,
            color="006100" if summary.net_result_eur >= 0 else "9C0006",
        )

        # ---- Pestaña 4: Instrucciones ----
        ws4 = wb.create_sheet("Instrucciones")
        ws4.column_dimensions["A"].width = 100
        instructions = [
            ("CÓMO USAR ESTE INFORME EN EL MODELO 100 — IRPF " + str(summary.year + 1), ""),
            ("", ""),
            ("1. GANANCIAS Y PÉRDIDAS PATRIMONIALES", ""),
            ("   Casilla 0389: Transmisiones de elementos patrimoniales (criptomonedas)", ""),
            ("   → Introduce el importe de 'Resultado neto base ahorro' de la pestaña Resumen Fiscal", ""),
            ("   → Este importe va en la BASE DEL AHORRO (no en base general)", ""),
            ("", ""),
            ("2. MÉTODO DE VALORACIÓN", ""),
            ("   El informe aplica FIFO (First In, First Out) que es el criterio obligatorio", ""),
            ("   según la consulta vinculante V0999-18 de la AEAT para criptomonedas.", ""),
            ("", ""),
            ("3. COMPENSACIÓN DE PÉRDIDAS", ""),
            ("   Las pérdidas de ejercicios anteriores (hasta 4 años) pueden compensarse.", ""),
            ("   Véase la pestaña Resumen para el importe de pérdidas arrastrables.", ""),
            ("", ""),
            ("4. DIRECTIVA DAC8", ""),
            ("   A partir del ejercicio 2026, los exchanges como OKX están obligados a", ""),
            ("   reportar automáticamente las operaciones a la AEAT (DAC8/DAC7 EU).", ""),
            ("   Los datos de la AEAT y este informe deben coincidir. En caso de discrepancia,", ""),
            ("   prevalece la información del exchange. Guarda todos los históricos de OKX.", ""),
            ("", ""),
            ("5. AVISO LEGAL", ""),
            ("   Este informe es orientativo. Consulta con un asesor fiscal para tu declaración.", ""),
            ("   Los tipos impositivos son los vigentes en 2026 (declaración IRPF 2025).", ""),
        ]
        ws4.append(["INSTRUCCIONES", ""])
        ws4[1][0].font = Font(bold=True, size=14, color="1F4E79")
        for row_data in instructions[1:]:
            ws4.append([row_data[0]])

        path = str(_REPORTS_DIR / f"informe_fiscal_{year}.xlsx")
        wb.save(path)
        logger.info("Excel guardado en {}", path)
        return path

    # -----------------------------------------------------------------------
    # JSON
    # -----------------------------------------------------------------------

    def _write_json(self, year: int, summary: FiscalSummary) -> str:
        data = {
            "year": summary.year,
            "total_trades": summary.total_trades,
            "total_gains": float(summary.total_gains_eur),
            "total_losses": float(summary.total_losses_eur),
            "net_result": float(summary.net_result_eur),
            "estimated_tax": float(summary.estimated_tax_eur),
            "deductible_fees": float(summary.deductible_fees_eur),
            "carryover_losses": float(summary.carryover_losses_eur),
            "usd_eur_rate": str(self._rate),
            "generated_at": summary.generated_at.isoformat(),
        }
        path = str(_REPORTS_DIR / f"informe_fiscal_{year}.json")
        Path(path).write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")
        logger.info("JSON guardado en {}", path)
        return path
